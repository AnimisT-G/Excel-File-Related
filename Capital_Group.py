from pathlib import Path
from os import system, listdir
import openpyxl as xl


def capital_group():
    print("""**Capital Group**\n
1. Report
2. Merge (2 Same Dimension Files)
3. Import (Old to New File)
""")
    option = input()
    if option == '1':
        report()
    elif option == '2':
        merge_2_excel_files()
    elif option == '3':
        import_old_to_new()
    else:
        input("\n**Wrong Input !!\n  Closing Program !!")


def report():
    file_name = excel_file_name_input(title='***CDM Report***\n')
    print(f'File - {file_name}\nOpening File...')

    workbook = xl.load_workbook(Path(file_name))
    sheet = workbook.active
    print("**File Opened Successfully !!!")

    # Printing Maximum Rows and Columns in the Sheet for proper Verification
    max_rows = sheet.max_row
    max_cols = sheet.max_column
    print(f"\nRows x Columns: {max_rows} x {max_cols}")

    # Generating the list of dates for which Report has to be Made
    date_input = input("\nEnter the start date(2022-07-13): ")
    check_date_format(date_input)

    print("\n**If output is needed for only one date then\n  just press 'Enter key'.")
    end_date_input = input("\nEnter the end date(2022-07-13): ")
    if end_date_input != '':
        check_date_format(end_date_input, date_input)

    dates = date_range_generator(date_input, end_date_input)

    # Determining the column number of Helper ID, Matching Status, Date Reviewed & Initials
    for i in range(1, max_cols + 1):
        if sheet.cell(1, i).value.lower() in ('helper', 'helper id', 'helper_id', 'helperid'):
            helper = i
        elif sheet.cell(1, i).value.lower() == 'disposition':
            m_nm = i - 1
            date_col = i + 4
            initials = i + 5
            break

    # Declaring the required Sets
    hid = set()  # "helperid" only
    mt_hid_st = set()  # "matchtype_helperid_sourcetype"
    total_reviewed_hid = set()  # "helperid_datereviewed"
    today_reviewed_hid = set()  # "herlperid" only
    today_data = set()  # "helperid_matching_initials_datereviewed" if on given dates

    # Filling the data from the sheet in above created Sets
    for i in range(2, max_rows + 1):
        # adding only Helper ID to hid
        hid.add(sheet.cell(i, helper).value)
        # adding Match_Type , Helper_ID and Source Type to mt_hid_st
        mt_hid_st.add((sheet.cell(i, helper - 1).value, sheet.cell(i,
                      helper).value, sheet.cell(i, helper + 1).value))
        # adding Helper ID & Dates (on which record Reviewed) to total_reviewed_hid
        total_reviewed_hid.add(
            (sheet.cell(i, helper).value, sheet.cell(i, date_col).value))

        if str(sheet.cell(i, date_col).value)[:10] in dates:
            # adding only Helper ID to today_reviewed_hid
            today_reviewed_hid.add(sheet.cell(i, helper).value)
            # adding Helper ID, Matching Status & Initials to today_data
            today_data.add((sheet.cell(i, helper).value, sheet.cell(
                i, m_nm).value, sheet.cell(i, initials).value, str(sheet.cell(i, date_col).value)[:10]))

    # Calculation the total reviewed and pending now records
    reviewed_now = []
    pending_now = []
    for y in total_reviewed_hid:
        if y[1] == None:
            pending_now.append(y[0])
        else:
            reviewed_now.append(y[0])

    # Calculating Matched and Not Matched Records and also counting the number of records done by each Member
    matched = 0
    not_matched = 0
    initial_counter = dict()  # Date+Inital : Value
    for d in today_data:
        if d[1].lower() == 'matched':
            matched += 1
        if d[1].lower() in ('not matched', 'not_matched'):
            not_matched += 1
        if d[3] + d[2] not in initial_counter:
            initial_counter[d[3] + d[2]] = 1
        else:
            initial_counter[d[3] + d[2]] += 1

    # Printing the total number of Unique Records
    print(f"\n\nTotal Unique Records (based on Helper ID): {len(hid)}")

    # Checking if Helper ID with different Match type and Source type exist
    if (len(hid) != len(mt_hid_st)):
        print(
            f"Total Unique Records (based on Column A,B,C): {len(mt_hid_st)}")
        print("\n**There are records with different 'Match Type' or 'Source Type'.")

    # Printing the Main Data for which Program has been Created
    print(f"""Statistics:
    Reviewed       : {len(today_reviewed_hid)} (based on Unique Helper Id's
    Matched        : {matched}
    Not Matched    : {not_matched}
    Pending        : {len(pending_now)}
    Total Reviewed : {len(reviewed_now)}
    """)

    # Checking if Matched & Not Matched data has any common Helper ID
    if len(today_data) != len(today_reviewed_hid):
        hid_in_m_nm = []
        for x in today_reviewed_hid:
            count = 0
            for y in today_data:
                if x == y[0]:
                    count += 1
            if count > 1:
                hid_in_m_nm.append(x)
        print("\n**Records in Matched & Not Matched with same Helper ID.")
        print(f"  Helper ID's - {hid_in_m_nm}")

    # Checking if the Reviewed & Pending records till now have common Helper ID
    if len(pending_now) + len(reviewed_now) != len(hid):
        hid_in_pnr = []
        for x in reviewed_now:
            if pending_now.count(x) > 0:
                hid_in_pnr.append(x)
        print("\n**Records in Reviewed & Pending with common Helper ID.")
        print(f"  Helper ID's - {hid_in_pnr}")

    # Printing the No. of records reviewed by each Member
    if dates[0] != dates[-1]:
        print(f"\nReviewed from {dates[0]} to {dates[-1]}:\n", end='')
    else:
        print(f"\nReviewed on {dates[0]}:\n", end='')

    unique_initials = sorted({key[10:] for key in sorted(initial_counter)})
    for date in dates:
        for initial in unique_initials:
            if date+initial not in initial_counter:
                initial_counter[date+initial] = 0

    initial_counter = refined_initialcounter(
        initial_counter, unique_initials, dates)

    sorted_keys_in_initial_counter = sorted(initial_counter)
    print("  Date         ", end='')
    for key in unique_initials:
        print(f'{key}   ', end='')
    old_date = ''
    total = 'Total'
    for key in sorted_keys_in_initial_counter:
        if old_date != key[:10]:
            print(f"{total}\n  {key[:10]} : %5d     " %
                  initial_counter[key], end='')
            old_date = key[:10]
            total = initial_counter[key]
        else:
            print("%5d     " % initial_counter[key], end='')
            total += initial_counter[key]
    print(f"{total}\n  Total      : ", end='')
    for key in unique_initials:
        total = 0
        for key2, value in initial_counter.items():
            if key2[10:] == key:
                total += value
        print('%5d     ' % total, end='')

    workbook.close()
    input("\n**File Closed Successfully !!!\n")


def merge_2_excel_files():
    file1 = excel_file_name_input(
        '1st', title='***Merge 2 Excel Files into 1st File***\n')
    file2 = excel_file_name_input(
        '2nd', title='***Merge 2 Excel Files into 1st File***\n')
    print(f'Filling  In  - {file1}\nFilling From - {file2}\nOpening Files...')

    workbook1 = xl.load_workbook(Path(file1))
    workbook2 = xl.load_workbook(Path(file2))
    sheet1 = workbook1.active
    sheet2 = workbook2.active

    max_rows1, max_cols1 = sheet1.max_row, sheet1.max_column
    max_rows2, max_cols2 = sheet2.max_row, sheet2.max_column

    if max_rows1 != max_rows2 or max_cols1 != max_cols2:
        print(f"\nFile1 -> Rows x Columns: {max_rows1} x {max_cols1}")
        print(f"File2 -> Rows x Columns: {max_rows2} x {max_cols2}")
        input("\n**Both Files doesn't have equal Rows and Columns.\n  Please Check the files.\n\n  Closing Program!!")
        exit()
    else:
        print(f"\nRows x Columns: {max_rows2} x {max_cols2}")

    flag = False
    cells_with_different_data = []
    for i in range(1, max_rows1 + 1):
        for j in range(1, max_cols1 + 1):
            if sheet1.cell(i, j).value == sheet2.cell(i, j).value:
                pass
            elif sheet1.cell(i, j).value != None and sheet2.cell(i, j).value == None:
                pass
            elif sheet1.cell(i, j).value == None and sheet2.cell(i, j).value != None:
                sheet1.cell(i, j).value = sheet2.cell(i, j).value
            elif sheet1.cell(i, j).value != None and sheet2.cell(i, j).value != None and sheet2.cell(i, j).value != sheet1.cell(i, j).value:
                flag = True
                cells_with_different_data.append(num_to_col_letter(j) + str(i))

    if flag == True:
        print(f"\n\n***There were Cells with different data in both files.")
        print("   So, Value in 1st File or Sheet is Considered.\n")
        print(cells_with_different_data)

    print("\n\n*** Merged Successfully!!! ***\n\nNew File created as 'Merge Result File.xlsx'\n\nSaving File...")
    workbook1.save("./Merge Result File.xlsx")
    input("\n*** File Saved !!! ***")


def import_old_to_new():
    old = excel_file_name_input(
        'Old', title='***Import Data from Old into New File***\n')
    new = excel_file_name_input(
        'New', title='***Import Data from Old into New File***\n')
    print(f'1. Old - {old}\n2. New - {new}\nOpening Files...')

    workbook1 = xl.load_workbook(Path(old))
    workbook2 = xl.load_workbook(Path(new))
    sheet1 = workbook1.active
    sheet2 = workbook2.active

    max_rows1, max_cols1 = sheet1.max_row, sheet1.max_column
    max_rows2, max_cols2 = sheet2.max_row, sheet2.max_column
    print(f"\nOld File -> Rows x Columns: {max_rows1} x {max_cols1}")
    print(f"New File -> Rows x Columns: {max_rows2} x {max_cols2}")

    # Determining the column numbers
    for i in range(1, max_cols1 + 1):
        if sheet1.cell(1, i).value.lower() in ('source plan sponsor name', 'source_plan_sponsor_name'):
            sheet1_sponsor_name = i
        elif sheet1.cell(1, i).value.lower() in ('plan name', 'plan_name'):
            sheet1_plan_name = i
        elif sheet1.cell(1, i).value.lower() in ('parsed org name reg line', 'parsed_org_name_reg_line'):
            sheet1_parsed_org_name = i
        elif sheet1.cell(1, i).value.lower() in ('matching cdm party id', 'matching_cdm_party_id', 'matching cdm partyid', 'matching_cdm_partyid'):
            sheet1_cdm = i
        elif sheet1.cell(1, i).value.lower() == 'disposition':
            sheet1_m_nm = i - 1
            sheet1_disposition = i
            sheet1_party = i + 2
            sheet1_plan = i + 3
            sheet1_date = i + 4
            sheet1_initial = i + 5
            sheet1_comment = i + 6
            break

    for i in range(1, max_cols2 + 1):
        if sheet2.cell(1, i).value.lower() in ('source plan sponsor name', 'source_plan_sponsor_name'):
            sheet2_sponsor_name = i
        elif sheet2.cell(1, i).value.lower() in ('plan name', 'plan_name'):
            sheet2_plan_name = i
        elif sheet2.cell(1, i).value.lower() in ('parsed org name reg line', 'parsed_org_name_reg_line'):
            sheet2_parsed_org_name = i
        elif sheet2.cell(1, i).value.lower() in ('matching cdm party id', 'matching_cdm_party_id', 'matching cdm partyid', 'matching_cdm_partyid'):
            sheet2_cdm = i
        elif sheet2.cell(1, i).value.lower() == 'disposition':
            sheet2_m_nm = i - 1
            sheet2_disposition = i
            sheet2_party = i + 2
            sheet2_plan = i + 3
            sheet2_date = i + 4
            sheet2_initial = i + 5
            sheet2_comment = i + 6
            break

    old_unique_data = []
    old_data = []

    print("\n**Unique Data considered on the basis of following columns:")
    print("  Sponsor Name, Plan Name, Parsed Org Name & CDM Party ID")

    print("\nLoading Data into Memory...")
    for i in range(2, max_rows1 + 1):
        if sheet1.cell(i, sheet1_initial) != None:
            temp = str(sheet1.cell(i, sheet1_sponsor_name).value) + str(sheet1.cell(i, sheet1_plan_name).value) + \
                str(sheet1.cell(i, sheet1_parsed_org_name).value) + \
                str(sheet1.cell(i, sheet1_cdm).value)
            if temp[:12] != 'NoneNoneNone':
                old_unique_data.append(temp.lower())
                temp = [temp.lower()]
                temp.append(sheet1.cell(i, sheet1_m_nm).value)
                temp.append(sheet1.cell(i, sheet1_disposition).value)
                temp.append(sheet1.cell(i, sheet1_party).value)
                temp.append(sheet1.cell(i, sheet1_plan).value)
                temp.append(sheet1.cell(i, sheet1_date).value)
                temp.append(sheet1.cell(i, sheet1_initial).value)
                temp.append(sheet1.cell(i, sheet1_comment).value)
                old_data.append(temp)

    print(f"\nLoading Data into New File from Memory...")
    for i in range(2, max_rows2 + 1):
        temp = str(sheet2.cell(i, sheet2_sponsor_name).value) + str(sheet2.cell(i, sheet2_plan_name).value) + \
            str(sheet2.cell(i, sheet2_parsed_org_name).value) + \
            str(sheet2.cell(i, sheet2_cdm).value)
        if temp.lower() in old_unique_data and temp[:12] != 'NoneNoneNone':
            index = old_unique_data.index(temp.lower())
            if old_data[index][0] == temp.lower():
                sheet2.cell(i, sheet2_m_nm).value = old_data[index][1]
                sheet2.cell(i, sheet2_disposition).value = old_data[index][2]
                sheet2.cell(i, sheet2_party).value = old_data[index][3]
                sheet2.cell(i, sheet2_plan).value = old_data[index][4]
                sheet2.cell(i, sheet2_date).value = old_data[index][5]
                sheet2.cell(i, sheet2_initial).value = old_data[index][6]
                sheet2.cell(i, sheet2_comment).value = old_data[index][7]

    print("\n\n*** Merged Successfully !!! ***\n\n*But need to fill ML Disposition by User itself.\n\nSaving File...")
    workbook2.save("./Imported to New File.xlsx")
    input("\n*** File Saved !!! ***")


def date_range_generator(start, end):
    dates = [start]
    if (end != start) and (end != ''):
        while start != end:
            start = increase_date(start)
            dates.append(start)
    return dates


def check_date_format(date1, small_date=None):
    days_in_months = [0, 31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    year, month, day = int(date1[:4]), int(date1[5:7]), int(date1[8:])
    if leap(year):
        days_in_months[2] = 29
    if (day > days_in_months[month]) or (12 < month < 1) or (date1[4] != '-') or (date1[4] != date1[7]) or (len(date1) != 10):
        print("\n\n**Wrong Date Format given. Please Re-Run program.\n")
        input()
        exit()
    if small_date != None and small_date != date1:
        syear, smonth, sday = int(small_date[:4]), int(
            small_date[5:7]), int(small_date[8:])
        if (syear > year) or (syear == year and smonth > month) or (syear == year and smonth == month and sday > day):
            print("\n\n**End Date can't come before Start Date in real world.\n")
            input()
            exit()

    return None


def increase_date(date1):
    days_in_months = [0, 31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    year, month, day = int(date1[:4]), int(date1[5:7]), int(date1[8:])
    if leap(year):
        days_in_months[2] = 29
    day += 1
    if day > days_in_months[month]:
        day, month = 1, month + 1

    if month == 13:
        month, year = 1, year + 1

    if year < 1000:
        year = f"0{year}"
    if month < 10:
        month = f"0{month}"
    if day < 10:
        day = f"0{day}"

    return f"{year}-{month}-{day}"


def leap(year):
    if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
        return True
    return False


def excel_file_name_input(string='', title=''):
    system('cls')
    if string != '':
        string += ' '
    print(title)
    print("Excel files in same directory:")
    files = listdir('.')
    xl_files = []
    counter = 0
    for file in files:
        if file[-4:] == 'xlsx':
            counter += 1
            xl_files.append(file)
            print(f"{counter}. {file}")

    if counter == 0:
        print("\n\n**No Files !!!\n  Closing Program !!")
        exit()
    elif counter == 1:
        return xl_files[0]
    option = int(
        input(f"\nEnter the file number corresponding to {string}file name: ")) - 1
    if len(xl_files) > option >= 0:
        system('cls')
        print(title)
        return xl_files[option]
    else:
        input("\n\n**Wrong Input !!!\n  Closing Program !!")
        exit()


def num_to_col_letter(num):
    letters = ''
    while num:
        mod = (num - 1) % 26
        letters += chr(mod + 65)
        num = (num - 1) // 26
    return ''.join(reversed(letters))


def refined_initialcounter(initial_counter, unique_initials, dates):
    for date in dates:
        check_zero = 0
        for initial in unique_initials:
            if initial_counter[date+initial] != 0:
                check_zero = 1
                break
        if check_zero == 0:
            for initial in unique_initials:
                del initial_counter[date+initial]

    return initial_counter


capital_group()
