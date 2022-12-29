from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException
from os import listdir
from pathlib import Path
from time import sleep
import openpyxl as xl

column_filter_flag = True
default_removals = "inc:inc.:,inc:llc:llc.:,llc:llp:llp.:,llp:co:co.:pa:p.a:pc:p.c:pllc:cpas:plan:and:&:trust:prof:ira:ltd:,ltd:ltd.:the:401:401k:401(k):k:(k):assetmark"


class BROWSER():
    def __init__(self):  # Opening Automated Webpage in Browser
        self.browser = webdriver.Edge()
        self.browser.get("https://cap.brightscope.com/search/beacon/#/plan")

    def sign_in(self, email, password):  # Sign In into Beacon
        self.browser.find_element(By.NAME, "email").click()
        action = ActionChains(self.browser)
        action.key_down(Keys.CONTROL).send_keys(
            'A').key_up(Keys.CONTROL).perform()
        self.browser.find_element(By.NAME, "email").send_keys(email)
        self.browser.find_element(By.NAME, "password").send_keys(password)
        # self.browser.find_element(By.TAG_NAME, "button").click()
        while True:
            try:
                self.browser.find_element(
                    By.CSS_SELECTOR, "button.main-search-button").click()
                return None
            except (NoSuchElementException, ElementNotInteractableException):
                try:
                    self.browser.find_element(By.CSS_SELECTOR, "ul.errors")

                    email, password = email_n_password()
                    self.sign_in(email, password)
                    return None
                except NoSuchElementException:
                    pass

    def search_filters(self):  # Activate Plan Name in Search Filter Navigation
        flag = False
        plan_name_checkbox = self.browser.find_element(
            By.XPATH, '//*[@id="filter-nav"]/div/div[2]/v-accordion/v-pane[2]/v-pane-content/div/div[14]/bale-filter/div/div[1]/md-checkbox/div[2]')
        while not flag:
            flag = self.browser.find_element(
                By.ID, "filter-nav").is_displayed()
            if flag:
                self.browser.find_elements(By.TAG_NAME, "v-pane")[1].click()
                flag = False
                while not flag:
                    flag = plan_name_checkbox.is_displayed()
        plan_name_checkbox.click()

    def column_filters(self):  # Activate EIN & Company Name in Column Filter Navigation
        self.browser.find_element(By.CSS_SELECTOR, "div.col-button").click()
        column_filter_navigation = self.browser.find_element(
            By.CSS_SELECTOR, "md-sidenav.md-sidenav-right")
        flag = False
        while not flag:
            flag = self.browser.find_element(
                By.XPATH, '//*[@id="body-content"]/ui-view/beacon-search/bale-search/bale-column-menu/md-sidenav/div/div[2]/v-accordion/v-pane[2]/v-pane-header/div/column-category/div/div[1]').is_displayed()
            if flag:
                self.browser.find_element(
                    By.XPATH, '//*[@id="body-content"]/ui-view/beacon-search/bale-search/bale-column-menu/md-sidenav/div/div[2]/v-accordion/v-pane[2]/v-pane-header/div/column-category/div/div[1]').click()
                flag = False
                while not flag:
                    flag = self.browser.find_element(
                        By.XPATH, '//*[@id="body-content"]/ui-view/beacon-search/bale-search/bale-column-menu/md-sidenav/div/div[2]/v-accordion/v-pane[2]/v-pane-content/div/div[11]/div/md-checkbox').is_displayed()
        self.browser.find_element(
            By.XPATH, '//*[@id="body-content"]/ui-view/beacon-search/bale-search/bale-column-menu/md-sidenav/div/div[2]/v-accordion/v-pane[2]/v-pane-content/div/div[11]/div/md-checkbox/div[2]').click()
        self.browser.find_element(
            By.XPATH, '//*[@id="body-content"]/ui-view/beacon-search/bale-search/bale-column-menu/md-sidenav/div/div[2]/v-accordion/v-pane[2]/v-pane-content/div/div[13]/div/md-checkbox/div[2]').click()
        try:
            send_us_email = self.browser.find_element(
                By.CSS_SELECTOR, "div.olark-text-button")
            if self.browser.find_element(By.CSS_SELECTOR, "div.olark-text-button").is_displayed():
                self.browser.execute_script("""
                    var element = arguments[0];
                    element.parentNode.removeChild(element);
                    """, send_us_email)
        except:
            pass

        self.browser.find_element(
            By.XPATH, '//*[@id="body-content"]/ui-view/beacon-search/bale-search/bale-column-menu/md-sidenav/div/div[3]/button/div[1]').click()
        flag = True
        while flag:
            flag = self.browser.find_element(
                By.CSS_SELECTOR, "md-sidenav.md-sidenav-right").is_displayed()

    def plan_name_search(self, pname):  # Type Plan Name in Field and Run Search
        flag = self.browser.find_element(By.ID, "filter-nav").is_displayed()
        if not flag:
            try:
                self.browser.find_element(
                    By.CSS_SELECTOR, "div.company-information").click()
            except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException):
                self.browser.find_element(
                    By.CSS_SELECTOR, "button.main-search-button").click()
        while True:
            try:
                self.browser.find_element(By.ID, "fl-input-125").click()
                break
            except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException):
                pass
        plan_name_input_field = self.browser.find_element(
            By.ID, "fl-input-125")
        action = ActionChains(self.browser)
        action.key_down(Keys.CONTROL).send_keys(
            'A').key_up(Keys.CONTROL).perform()
        plan_name_input_field.send_keys(pname)
        flag = False
        plan_name_field_list = self.browser.find_element(
            By.XPATH, '//*[@id="ul-125"]')
        sleep(1)
        while True:
            try:
                sleep(1)
                flag = plan_name_field_list.is_displayed()
                if flag:
                    self.browser.find_element(
                        By.XPATH, '//*[@id="ul-125"]/li[1]/md-autocomplete-parent-scope').click()
                    break
                else:
                    plan_name_input_field.click()
            except (NoSuchElementException, ElementNotInteractableException):
                pass
        self.browser.find_element(
            By.CSS_SELECTOR, "button.submit-button").click()

    def results(self):  # Search Results
        flag = True
        while flag:
            flag = self.browser.find_element(
                By.ID, "filter-nav").is_displayed()
        flag = False
        while not flag:
            try:
                flag = self.browser.find_element(
                    By.CSS_SELECTOR, "span.small-header").is_displayed()
                global column_filter_flag
                if column_filter_flag:
                    self.column_filters()
                    column_filter_flag = False
            except NoSuchElementException:
                try:
                    self.browser.find_element(
                        By.CSS_SELECTOR, "button.main-search-button").click()
                    return 0, self.browser.find_elements(By.CSS_SELECTOR, "div.search-apology")
                except (NoSuchElementException, ElementClickInterceptedException):
                    pass
        search_results = self.browser.find_element(
            By.CSS_SELECTOR, "span.small-header").text
        results_plan = self.browser.find_elements(
            By.CSS_SELECTOR, "div.result-cell")
        search_results = search_results[16:-1]
        if len(search_results) > 2:
            search_results = 30
        return int(search_results), results_plan

    def quit(self):  # Close the Automated Webpage
        self.browser.quit()


def main():  # Main Function
    # Email ID and Password for LogIn into Beacon
    global removals
    try:
        with open("login.txt", 'r') as file:
            lines = file.readlines()
            email, password, removals = lines[0], lines[1], lines[2].split(':')
    except FileNotFoundError:
        email, password = email_n_password()
        removals = default_removals.split(':')

    file_name = excel_file_name_input()
    print(f'File - {file_name}\nOpening File...')

    data_excel = xl.load_workbook(Path(file_name))
    data_sheet = data_excel.active
    print("File Opened Successfully!!")
    new_excel = xl.Workbook()
    sheet = new_excel.active
    sheet[f"A1"].value = data_sheet[f"B1"].value
    sheet[f"B1"].value = data_sheet[f"Y1"].value
    sheet[f"C1"].value = "Cleaned Org Name"

    # Creating the Web Automation Object
    driver = BROWSER()
    driver.sign_in(email, password)
    driver.search_filters()

    previous_plan = previous_cleaned_plan = 'a'

    for row in range(2, data_sheet.max_row):
        date = data_sheet[f"AP{row}"].value  # date
        if date != None:
            continue

        sheet[f"A{row}"].value = data_sheet[f"B{row}"].value  # helper_id
        sheet[f"B{row}"].value = data_sheet[f"Y{row}"].value  # parsed_org_name
        plan = str(sheet[f"B{row}"].value)
        cleaned_plan = check_plan(plan)
        if (plan == previous_plan) or (cleaned_plan == previous_cleaned_plan):
            sheet[f"D{row}"].value = sheet[f"D{row - 1}"].value
            continue
        elif len(cleaned_plan) < 5:
            sheet[f"D{row}"].value = 'Not Searched'
            continue

        previous_plan = plan
        sheet[f"C{row}"].value = previous_cleaned_plan = cleaned_plan
        driver.plan_name_search(cleaned_plan)
        new_excel.save(f"./Automated Results.xlsx")
        search_results, results_plan = driver.results()
        if search_results == 0:
            sheet[f"D{row}"].value = 'Not Found'
            continue
        elif search_results > 29:
            sheet[f"D{row}"].value = 'Multiple Webpages'
            continue

        Results = {'Plans': [], 'Dates': [],
                   'Company': [], 'EIN': [], 'FA': []}
        for i in range(search_results):
            Results['Plans'].append(results_plan[5+i].text)
            Results['Dates'].append(results_plan[5+i+search_results].text)
            Results['Company'].append(results_plan[5+i+search_results*2].text)
            Results['EIN'].append(results_plan[5+i+search_results*3].text)
            Results['FA'].append(results_plan[5+i+search_results*4].text)

        sheet[f"D{row}"].value = decision(
            search_results, Results, cleaned_plan)
    driver.quit()

    new_excel.save(f"./Automated Results.xlsx")
    new_excel.close()


def decision(search_results, Results, cleaned_plan):
    count = [0, [], []]
    for i in range(search_results):
        cleaned_company = check_plan(Results['Company'][i])
        if cleaned_company == cleaned_plan:
            count[0] += 1
            count[1].append(i)
            count[2].append(Results['EIN'][i])

    ein = list(set(count[2]))
    if len(ein) == 1:
        if ein[0] == '':
            return "Beacon: Found without EIN"
        else:
            return ein[0]
    else:
        return "Mannual Check is Required"


def email_n_password():  # Asking User Email & Password if Unable to Login
    email = input("Enter Login Email    : ")
    password = input("Enter Login Password : ")
    option = input("\nSAVE Login Credentials in 'login.txt' file (Y) : ")
    if option in ('Y', 'y'):
        with open("login.txt", 'w') as file:
            file.write(email+'\n'+password+'\n'+default_removals)
    return email, password


def check_plan(plan=''):
    temp = plan.lower().split(' ')
    plan = ''
    for word in temp:
        if (word in removals) or (len(word) < 3):
            continue
        else:
            c = word.find("'s")
            if c > 0:
                word = word[:c]
            r = [',', '.']
            word = list(word)
            for j in r:
                if j in word:
                    word.remove(j)
            plan += ''.join(word) + ' '
    return plan.strip()


def excel_file_name_input():
    print("Excel files in same directory:")
    files = listdir('.')
    xl_files = []
    counter = 0
    for file in files:
        if file[-4:] == 'xlsx':
            counter += 1
            xl_files.append(file)
            print(f"{counter}. {file}")
    if counter == 1:
        return xl_files[0]
    elif counter == 0:
        input("\n**No Excel Files Found.\n  Closing Program !!")
        exit()
    option = int(
        input(f"\nEnter the file number corresponding to file name: ")) - 1
    if len(xl_files) > option >= 0:
        return xl_files[option]
    else:
        input("\n\n**Wrong Input !!!\n  Closing Program !!")
        exit()


main()
input("Program Finished!!")
